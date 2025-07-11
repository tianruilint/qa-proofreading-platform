import json
import os
from datetime import datetime
from openpyxl import Workbook
from werkzeug.utils import secure_filename
from src.config import Config

class FileHandler:
    @staticmethod
    def validate_jsonl_file(file_path):
        """验证JSONL文件格式"""
        try:
            qa_pairs = []
            with open(file_path, 'r', encoding='utf-8') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line:
                        continue
                    
                    try:
                        data = json.loads(line)
                        if not isinstance(data, dict):
                            return False, f"第{line_num}行不是有效的JSON对象"
                        
                        if 'prompt' not in data or 'completion' not in data:
                            return False, f"第{line_num}行缺少必需的字段 'prompt' 或 'completion'"
                        
                        if not isinstance(data['prompt'], str) or not isinstance(data['completion'], str):
                            return False, f"第{line_num}行的 'prompt' 和 'completion' 必须是字符串"
                        
                        qa_pairs.append({
                            'prompt': data['prompt'],
                            'completion': data['completion'],
                            'original_index': line_num - 1
                        })
                        
                    except json.JSONDecodeError as e:
                        return False, f"第{line_num}行JSON格式错误: {str(e)}"
            
            if len(qa_pairs) == 0:
                return False, "文件为空或没有有效的QA对"
            
            if len(qa_pairs) > Config.MAX_QA_PAIRS_PER_FILE:
                return False, f"文件包含{len(qa_pairs)}个QA对，超过最大限制{Config.MAX_QA_PAIRS_PER_FILE}"
            
            return True, qa_pairs
            
        except Exception as e:
            return False, f"文件读取错误: {str(e)}"
    
    @staticmethod
    def save_uploaded_file(file, folder):
        """保存上传的文件"""
        if not os.path.exists(folder):
            os.makedirs(folder)
        
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        name, ext = os.path.splitext(filename)
        unique_filename = f"{name}_{timestamp}{ext}"
        
        file_path = os.path.join(folder, unique_filename)
        file.save(file_path)
        
        return file_path, unique_filename
    
    @staticmethod
    def export_to_jsonl(qa_pairs, output_path):
        """导出QA对为JSONL格式"""
        try:
            with open(output_path, 'w', encoding='utf-8') as f:
                for qa_pair in qa_pairs:
                    if not qa_pair.get('is_deleted', False):
                        json_line = json.dumps({
                            'prompt': qa_pair['prompt'],
                            'completion': qa_pair['completion']
                        }, ensure_ascii=False)
                        f.write(json_line + '\n')
            return True, None
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    def export_to_excel(qa_pairs, output_path):
        """导出QA对为Excel格式"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "QA对"
            
            # 设置表头
            headers = ['序号', '问题(prompt)', '答案(completion)']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 写入数据
            row = 2
            for index, qa_pair in enumerate(qa_pairs, 1):
                if not qa_pair.get('is_deleted', False):
                    ws.cell(row=row, column=1, value=index)
                    ws.cell(row=row, column=2, value=qa_pair['prompt'])
                    ws.cell(row=row, column=3, value=qa_pair['completion'])
                    row += 1
            
            # 调整列宽
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 50
            ws.column_dimensions['C'].width = 50
            
            wb.save(output_path)
            return True, None
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    def generate_export_filename(original_filename, export_type='jsonl'):
        """生成导出文件名"""
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        name, _ = os.path.splitext(original_filename)
        
        if export_type == 'jsonl':
            return f"{name}_edited_{timestamp}.jsonl"
        elif export_type == 'excel':
            return f"{name}_{timestamp}.xlsx"
        else:
            return f"{name}_{timestamp}.{export_type}"
    
    # 移除自动删除功能 - 文件将永久保留，除非用户主动删除
    # @staticmethod
    # def cleanup_expired_files():
    #     """清理过期文件 - 此功能已被禁用"""
    #     pass

