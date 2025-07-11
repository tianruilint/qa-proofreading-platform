import os
import json
import uuid
from datetime import datetime
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from flask import current_app

def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in current_app.config['ALLOWED_EXTENSIONS']

def generate_unique_filename(original_filename):
    """生成唯一的文件名"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    unique_id = str(uuid.uuid4())[:8]
    name, ext = os.path.splitext(secure_filename(original_filename))
    return f"{name}_{timestamp}_{unique_id}{ext}"

def save_uploaded_file(file, upload_folder):
    """保存上传的文件"""
    if not file or not allowed_file(file.filename):
        return None, "不支持的文件类型"
    
    filename = generate_unique_filename(file.filename)
    file_path = os.path.join(upload_folder, filename)
    
    try:
        file.save(file_path)
        file_size = os.path.getsize(file_path)
        return {
            'filename': filename,
            'original_filename': file.filename,
            'file_path': file_path,
            'file_size': file_size,
            'file_type': file.filename.rsplit('.', 1)[1].lower()
        }, None
    except Exception as e:
        return None, f"文件保存失败: {str(e)}"

def parse_jsonl_file(file_path):
    """解析JSONL文件"""
    try:
        qa_pairs = []
        with open(file_path, 'r', encoding='utf-8') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line:
                    continue
                
                try:
                    data = json.loads(line)
                    if not isinstance(data, dict):
                        return None, f"第{line_num}行不是有效的JSON对象"
                    
                    if 'prompt' not in data or 'completion' not in data:
                        return None, f"第{line_num}行缺少必需的字段 'prompt' 或 'completion'"
                    
                    qa_pairs.append({
                        'prompt': str(data['prompt']),
                        'completion': str(data['completion'])
                    })
                
                except json.JSONDecodeError as e:
                    return None, f"第{line_num}行JSON格式错误: {str(e)}"
        
        if not qa_pairs:
            return None, "文件中没有有效的QA对"
        
        return qa_pairs, None
    
    except Exception as e:
        return None, f"文件读取失败: {str(e)}"

def export_to_jsonl(qa_pairs, output_path):
    """导出QA对到JSONL文件"""
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            for qa_pair in qa_pairs:
                json_line = json.dumps({
                    'prompt': qa_pair['prompt'],
                    'completion': qa_pair['completion']
                }, ensure_ascii=False)
                f.write(json_line + '\n')
        
        return True, None
    except Exception as e:
        return False, f"导出JSONL文件失败: {str(e)}"

def export_to_excel(qa_pairs, output_path):
    """导出QA对到Excel文件"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "QA对"
        
        # 设置表头
        headers = ['序号', '问题(prompt)', '答案(completion)']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # 写入数据
        for row, qa_pair in enumerate(qa_pairs, 2):
            ws.cell(row=row, column=1, value=row - 1)  # 序号
            ws.cell(row=row, column=2, value=qa_pair['prompt'])
            ws.cell(row=row, column=3, value=qa_pair['completion'])
        
        # 调整列宽
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 50
        
        wb.save(output_path)
        return True, None
    
    except Exception as e:
        return False, f"导出Excel文件失败: {str(e)}"

def create_export_filename(original_filename, export_type='jsonl', suffix='edited'):
    """创建导出文件名"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    name, _ = os.path.splitext(original_filename)
    
    if export_type == 'excel':
        return f"{name}_{suffix}_{timestamp}.xlsx"
    else:
        return f"{name}_{suffix}_{timestamp}.jsonl"

def get_file_info(file_path):
    """获取文件信息"""
    try:
        stat = os.stat(file_path)
        return {
            'size': stat.st_size,
            'created_at': datetime.fromtimestamp(stat.st_ctime),
            'modified_at': datetime.fromtimestamp(stat.st_mtime),
            'exists': True
        }
    except OSError:
        return {
            'exists': False
        }

def delete_file_safe(file_path):
    """安全删除文件"""
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            return True, None
    except Exception as e:
        return False, f"删除文件失败: {str(e)}"
    
    return True, None

def ensure_directory_exists(directory):
    """确保目录存在"""
    try:
        os.makedirs(directory, exist_ok=True)
        return True
    except Exception:
        return False

def validate_jsonl_content(content):
    """验证JSONL内容格式"""
    if not isinstance(content, list):
        return False, "内容必须是数组格式"
    
    for index, item in enumerate(content):
        if not isinstance(item, dict):
            return False, f"第{index + 1}项不是有效的对象"
        
        if 'prompt' not in item or 'completion' not in item:
            return False, f"第{index + 1}项缺少必需的字段 'prompt' 或 'completion'"
        
        if not isinstance(item['prompt'], str) or not isinstance(item['completion'], str):
            return False, f"第{index + 1}项的 'prompt' 和 'completion' 必须是字符串"
    
    return True, None

def split_qa_pairs_for_assignment(qa_pairs, assignments):
    """为任务分配拆分QA对"""
    """
    assignments: [
        {'user_id': 1, 'count': 10},
        {'user_id': 2, 'count': 15},
        ...
    ]
    """
    total_pairs = len(qa_pairs)
    total_assigned = sum(assignment['count'] for assignment in assignments)
    
    if total_assigned > total_pairs:
        return None, "分配的QA对数量超过了总数量"
    
    result = []
    start_index = 0
    
    for assignment in assignments:
        count = assignment['count']
        if count <= 0:
            continue
        
        end_index = start_index + count - 1
        result.append({
            'user_id': assignment['user_id'],
            'start_index': start_index,
            'end_index': end_index,
            'qa_pairs': qa_pairs[start_index:start_index + count]
        })
        
        start_index += count
    
    return result, None

def merge_qa_pairs_from_assignments(assignments_data):
    """合并来自不同分配的QA对"""
    """
    assignments_data: [
        {
            'start_index': 0,
            'end_index': 9,
            'qa_pairs': [...]
        },
        ...
    ]
    """
    # 按start_index排序
    assignments_data.sort(key=lambda x: x['start_index'])
    
    merged_pairs = []
    for assignment in assignments_data:
        merged_pairs.extend(assignment['qa_pairs'])
    
    return merged_pairs

